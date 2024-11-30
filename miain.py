import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os 

name_file = 'datos.xlsx'
if os.path.exists(name_file):
    wb = load_workbook(name_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(['Nombre', 'Edad', 'Email', 'Telefono', 'Dirección'])



#Cargar datos
def load():
    name = entry_name.get()
    age = entry_age.get()
    email = entry_email.get()
    phone = entry_phone.get()
    adress = entry_adress.get()

    if not name or not age or not email or not phone or not adress:
        messagebox.showerror('Advertencia', 'Todos los campos son obligatorios')
        return
    
    if not age.isdigit() or not phone.isdigit():
        messagebox.showerror("Advertencia", 'Edad y Telefono deben ser enteros.')
        return
    
    if '@' not in email:
        messagebox.showerror("Advertencia", 'El email debe contener "@".',)
        return
    
    ws.append([name, age, email, phone, adress])
    wb.save(name_file)
    messagebox.showinfo("Información", 'Datos cargados con éxito')

    entry_name.delete(0, tk.END)
    entry_age.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_phone.delete(0, tk.END)
    entry_adress.delete(0, tk.END)

    

root = tk.Tk()
root.title("Formulario de Datos")
root.configure(bg='#4B6587')
label_style = {'bg': '#4B6587', 'fg': 'white'}
entry_style = {'bg': '#D3D3D3', 'fg': 'black'}

#Nombre
label_name = tk.Label(root, text="Nombre", **label_style)
label_name.grid(row=0, column=0, padx=10, pady=5)
entry_name = tk.Entry(root, **entry_style)
entry_name.grid(row=0, column=1, padx=10, pady=5)

#Edad
label_age = tk.Label(root, text="Edad", **label_style)
label_age.grid(row=1, column=0, padx=10, pady=5)
entry_age = tk.Entry(root, **entry_style)
entry_age.grid(row=1, column=1, padx=10, pady=5)

#Email
label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

#Telefono
label_phone = tk.Label(root, text="Teléfono", **label_style)
label_phone.grid(row=3, column=0, padx=10, pady=5)
entry_phone = tk.Entry(root, **entry_style)
entry_phone.grid(row=3, column=1, padx=10, pady=5)

#Direccion
label_adress = tk.Label(root, text="Dirección", **label_style)
label_adress.grid(row=4, column=0, padx=10, pady=5)
entry_adress = tk.Entry(root, **entry_style)
entry_adress.grid(row=4, column=1, padx=10, pady=5)

button = tk.Button(root, text="Guardar", command=load, bg='#6D8299', fg='white')
button.grid(row=5, column=0, columnspan=2 ,padx=10, pady=10)

root.mainloop()